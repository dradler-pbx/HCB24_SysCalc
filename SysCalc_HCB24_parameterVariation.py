import ComponentsLibrary as cmp
from scipy.optimize import fsolve
import numpy as np
from CoolProp.CoolProp import PropsSI as CPPSI
from openpyxl import workbook, load_workbook
from tqdm import tqdm

wb = load_workbook("ParameterVariation.xlsx")
# check if write permission is granted
wb.save("ParameterVariation.xlsx")
sh = wb[wb.sheetnames[0]]
# kA_IHX = 2.3

for i in tqdm(range(3, sh.max_column+1)):
    data = []
    for row in range(2, 23):
        data.append(sh.cell(row=row, column=i).value)
    [Refrigerant, SecLiquid, kEVAP, AEVAP, dTSH, kCOND, ACOND, mACOND, dTSC, kHEX, AHEX, mABox, TBox, TAi, mSL, RPM, eta_V, eta_S, stroke, kA_IHX, tolerance] = data
    TBox = TBox + 273.15
    TAi = TAi + 273.15
    T0 = TBox - 10
    T7 = T0 + dTSH
    TC = TAi +10
    (TAoDesuperheat, TAoCondenser, TAoSubcool) = ((TC+TAi)/2, (TC+TAi)/2, (TC+TAi)/2)
    T1 = T7 + 1.
    T5 = TC - 1.

    p0 = CPPSI("P","T",T0,"Q",1,Refrigerant)
    pC = CPPSI("P","T",TC,"Q",0,Refrigerant)
    TSL1 = TBox -2
    TSL2 = TSL1 - 3
    TSLmid = TSL2 + 2.5
    (xC1,xC2,xC3) = (0.05,0.8,0.15)
    (xE1,xE2) = (0.8,0.2)
    TAHEX = TBox - 2

    '''Simulation Loop'''
    for iteration in range(1,1000):
        # print(iteration)
        p0prev = p0
        pCprev = pC
        TSL1prev = TSL1
        # print(iteration)
        # (mR, Pel, h2, T2) = cmp.Poly_CPR_varRPM(p0, pC, RPM)
        (mR, Pel, h2, T2) = cmp.CPR_efficiency_model(pin=p0, Tin=T1, pC=pC, RPM=RPM, eff_S=eta_S, eff_V=eta_V, dV=stroke)
        # print(p0, pC, RPM)
        # print(cmp.Poly_CPR_varRPM(p0, pC, RPM))
        # print([TC, TAoDesuperheat, TAoCondenser, TAoSubcool, xC1, xC2, xC3])
        # print([mR, mACOND, ACOND, T2, TAi, dTSC, kCOND])
        (TC, TAoDesuperheat, TAoCondenser, TAoSubcool, xC1, xC2, xC3) = fsolve(cmp.COND3Zone,[TC, TAoDesuperheat, TAoCondenser, TAoSubcool, xC1, xC2, xC3], [mR, mACOND, ACOND, T2, TAi, 0, kCOND])
        pC = CPPSI("P","T",TC,"Q",0,Refrigerant)
        h3 = CPPSI("H","P",pC,"Q",0,Refrigerant)
        T4 = TC
        (T5, T1) = fsolve(cmp.IHX, [T5, T1], [kA_IHX, T7, mR, mR, Refrigerant, T4, pC, p0])
        h5 = CPPSI('H', 'T', T5, 'P', pC, Refrigerant)
        (T0, TSL2, TSLmid, xE1, xE2) = fsolve(cmp.EVAP,[T0, TSL2, TSLmid, xE1, xE2], [mR, h5, mSL, TSL1, kEVAP, AEVAP, dTSH, SecLiquid, False])
        T7 = T0 + dTSH
        (T1, T5) = fsolve(cmp.IHX, [T1, T5], [kA_IHX, T4, mR, mR, Refrigerant, T7, p0, pC])
        p0 = CPPSI("P","T", T0, "Q", 1, Refrigerant)
        (TSL1, TAHEX)=fsolve(cmp.HeatEx, [TSL1, TAHEX], [mSL, TSL2, mABox, TBox, kHEX, AHEX, SecLiquid])

        if np.abs(p0-p0prev) < tolerance and np.abs(pC-pCprev) < tolerance and np.abs(TSL1-TSL1prev) < tolerance:
            break

    h7 = CPPSI('H', 'T', T7, 'P', p0, Refrigerant)
    h1 = CPPSI('H', 'T', T1, 'P', p0, Refrigerant)
    output_data = [' '. join([str(np.abs(p0-p0prev)), str(np.abs(pC-pCprev)), str(np.abs(TSL1-TSL1prev))]),
                   h1,
                   p0,
                   h2,
                   pC,
                   h3,
                   pC,
                   h3,
                   pC,
                   h5,
                   pC,
                   h5,
                   p0,
                   h7,
                   p0,
                   mR,
                   Pel,
                   TSL1-273.15,
                   TSL2-273.15,
                   TAoCondenser-273.15,
                   TAHEX-273.15]
    for row in range(23, 44):
        value = output_data.pop(0)
        sh.cell(row=row, column=i).value = value

wb.save("ParameterVariation.xlsx")
wb.close()
print('done!')




